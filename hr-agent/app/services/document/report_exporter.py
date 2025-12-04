"""Report Exporter - Generate PDF and Excel reports."""

import io
import logging
from datetime import date
from typing import Any, Dict, Optional

logger = logging.getLogger(__name__)


def generate_pdf_report(report_data: Dict[str, Any]) -> bytes:
    """
    Generate PDF report from report data.

    Args:
        report_data: Report data dictionary

    Returns:
        PDF content as bytes
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=30,
        )
        story.append(Paragraph(f"HR Report: {report_data.get('report_type', 'Unknown')}", title_style))
        story.append(Spacer(1, 12))

        # Metadata
        meta_style = styles["Normal"]
        story.append(Paragraph(f"Generated: {report_data.get('generated_at', date.today())}", meta_style))
        story.append(Paragraph(f"Report ID: {report_data.get('id', 'N/A')}", meta_style))
        story.append(Spacer(1, 24))

        # Report content
        data = report_data.get("data", {})

        if report_data.get("report_type") == "headcount":
            # Headcount summary
            story.append(Paragraph("Summary", styles["Heading2"]))
            story.append(Spacer(1, 12))

            summary_data = [
                ["Metric", "Value"],
                ["Total Employees", str(data.get("total_employees", 0))],
                ["Active Employees", str(data.get("active_employees", 0))],
                ["New Hires (Month)", str(data.get("new_hires_this_month", 0))],
                ["Terminations (Month)", str(data.get("terminations_this_month", 0))],
                ["Net Change", str(data.get("net_change", 0))],
            ]

            table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(table)
            story.append(Spacer(1, 24))

            # By department
            if data.get("by_department"):
                story.append(Paragraph("By Department", styles["Heading2"]))
                story.append(Spacer(1, 12))

                dept_data = [["Department", "Count"]]
                for dept in data["by_department"]:
                    dept_data.append([
                        dept.get("department_name", "Unknown"),
                        str(dept.get("employee_count", 0)),
                    ])

                table = Table(dept_data, colWidths=[4 * inch, 1.5 * inch])
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (1, 0), (1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(table)

        else:
            # Generic data display
            story.append(Paragraph("Report Data", styles["Heading2"]))
            story.append(Spacer(1, 12))
            story.append(Paragraph(str(data), styles["Normal"]))

        doc.build(story)
        return buffer.getvalue()

    except ImportError:
        logger.error("reportlab not installed. Cannot generate PDF reports.")
        # Return a simple text-based PDF alternative
        return b"%PDF-1.4 (Report generation requires reportlab library)"
    except Exception as e:
        logger.error(f"Failed to generate PDF: {e}")
        raise


def generate_excel_report(report_data: Dict[str, Any]) -> bytes:
    """
    Generate Excel report from report data.

    Args:
        report_data: Report data dictionary

    Returns:
        Excel content as bytes
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "HR Report"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center")

        # Title
        ws["A1"] = f"HR Report: {report_data.get('report_type', 'Unknown')}"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        ws["A2"] = f"Generated: {report_data.get('generated_at', date.today())}"
        ws["A3"] = f"Report ID: {report_data.get('id', 'N/A')}"

        current_row = 5
        data = report_data.get("data", {})

        if report_data.get("report_type") == "headcount":
            # Summary section
            ws.cell(row=current_row, column=1, value="Summary").font = Font(bold=True, size=14)
            current_row += 1

            summary_items = [
                ("Total Employees", data.get("total_employees", 0)),
                ("Active Employees", data.get("active_employees", 0)),
                ("New Hires (Month)", data.get("new_hires_this_month", 0)),
                ("Terminations (Month)", data.get("terminations_this_month", 0)),
                ("Net Change", data.get("net_change", 0)),
            ]

            for label, value in summary_items:
                ws.cell(row=current_row, column=1, value=label)
                ws.cell(row=current_row, column=2, value=value)
                current_row += 1

            current_row += 2

            # By department
            if data.get("by_department"):
                ws.cell(row=current_row, column=1, value="By Department").font = Font(bold=True, size=14)
                current_row += 1

                # Headers
                headers = ["Department", "Employee Count", "Manager"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                current_row += 1

                for dept in data["by_department"]:
                    ws.cell(row=current_row, column=1, value=dept.get("department_name", "Unknown"))
                    ws.cell(row=current_row, column=2, value=dept.get("employee_count", 0))
                    ws.cell(row=current_row, column=3, value=dept.get("manager_name", ""))
                    current_row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except ImportError:
        logger.error("openpyxl not installed. Cannot generate Excel reports.")
        return b"Excel generation requires openpyxl library"
    except Exception as e:
        logger.error(f"Failed to generate Excel: {e}")
        raise
